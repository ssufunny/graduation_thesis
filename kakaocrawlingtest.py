from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
import pandas as pd
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import ElementNotInteractableException

all_values = []
load_wb = load_workbook(filename="C://Users//정보통신공학과//Downloads//all_place.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

for r in range(1686,18570):
    row_value = ""
    for c in range(2,6):
        if load_ws.cell(row=r, column=c).value == None:
            row_value += ""
        else:
            row_value += str(load_ws.cell(row=r, column=c).value)+" "
    all_values.append(row_value)

driver = webdriver.Chrome("C://Users//정보통신공학과//Downloads//chromedriver_win32 (2)//chromedriver.exe")
driver.maximize_window()
driver.get("https://map.kakao.com/")
# 주소창 찾기
elem = driver.find_element(By.NAME, value="q")
f = open("상호명전국1686.txt", 'w', encoding='utf-8')
#12개, 0개, 4페이지까지, '끝까지'

#all_values = ['충북 청주시 청원구 내수읍 덕암 2길','충북 청주시 청원구 내수읍 덕암 2길', '충북 청주시 청원구 내수읍 덕암 24길','충북 청주시 청원구 내수읍 초정약수로','충북 청주시 청원구 내수읍']
for seoul_gu in all_values:
    elem.clear()
    elem.send_keys(seoul_gu)
    driver.find_element_by_xpath('//*[@id="search.keyword.submit"]').send_keys(Keys.ENTER)
    time.sleep(0.7)
    # 장소 더보기 누르기
    try: # 장소 더보기 누르기 ( 장소 있는 경우 )
        driver.find_element_by_css_selector("#info\.search\.place\.more").send_keys(Keys.ENTER)
        time.sleep(0.4)
        print(seoul_gu)

        while True:  # 페이지들 크롤링이 전부 끝날 때까지 계속 [다음] 버튼으로 넘어감

            for i in range(1, 6):
                # 한 덩어리에는 5개의 페이지가 존재 (1페이지 to 5페이지 / 6페이지 to 10페이지 / .. etc.)

                xPath_page = '//*[@id="info.search.page.no' + str(i) + '"]'
                try:
                    driver.find_element_by_xpath(xPath_page).send_keys(Keys.ENTER)
                    time.sleep(0.4)

                    shops = driver.find_elements_by_class_name("PlaceItem")

                    for shop in shops:
                        # 이름
                        name = shop.find_element_by_css_selector("div.head_item.clickArea > strong > a.link_name").text
                        # 별 리뷰
                        try:
                            rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                            rating_num = rating[:-1]
                            # 참조 리뷰
                            review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                            if int(rating_num) + int(review) > 1:
                                f.writelines([seoul_gu + ",", name + ",", rating_num + ",", review + "\n"])
                        except:
                            continue
                except ElementNotInteractableException:
                    if (i==1):
                        shops = driver.find_elements_by_class_name("PlaceItem")

                        for shop in shops:
                            # 이름
                            name = shop.find_element_by_css_selector(
                                "div.head_item.clickArea > strong > a.link_name").text
                            # 별 리뷰
                            try:
                                rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                                rating_num = rating[:-1]
                                # 참조 리뷰
                                review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                                if int(rating_num) + int(review) > 1:
                                    f.writelines([seoul_gu + ",", name + ",", rating_num + ",", review + "\n"])
                            except:
                                continue
                    else:
                        break

            # [다음] 버튼의 클래스 속성 값이 next 이면 계속 넘어가고, 아니면(next disabled) 크롤링 종료
            next_button = driver.find_element_by_xpath('//*[@id="info.search.page.next"]')
            next_button_class = next_button.get_attribute('class')

            if next_button_class == "next":
                xPath_next_button = '//*[@id="info.search.page.next"]'
                driver.find_element_by_xpath(xPath_next_button).send_keys(Keys.ENTER)
            else:
                break

    except: # 더보기 없으면 수집 -> 장소 있을수도, 없을수도 하여튼간 15개 이하
        try:
            shops = driver.find_elements_by_class_name("PlaceItem")


            for shop in shops:
                # 이름
                name = shop.find_element_by_css_selector("div.head_item.clickArea > strong > a.link_name").text

                # 별 리뷰
                try:
                    rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                    rating_num = rating[:-1]
                    # 참조 리뷰
                    review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                    if int(rating_num)+int(review) > 1:
                        f.writelines([seoul_gu+",", name+",", rating_num+",", review+"\n"])
                except:
                    continue
        except: #암것도 없으면
            continue
