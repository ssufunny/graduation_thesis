from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
import pandas as pd
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# 코딩 시작! --------------------------------------------------------------------------------------------

# 셀레니움에 쓸 크롬 드라이버 시작
driver = webdriver.Chrome("C://Users//정보통신공학과//Downloads//chromedriver_win32 (2)//chromedriver.exe")
# 카카오지도 들어가기
driver.maximize_window()
driver.get("https://map.kakao.com/")
# 주소창 찾기
elem = driver.find_element(By.NAME, value="q")

#search_list = pd.read_excel("C://Users//정보통신공학과//Downloads//place.xlsx", usecols=[1,2,3,4])
all_values = []
load_wb = load_workbook(filename="C://Users//정보통신공학과//Downloads//place.xlsx", data_only=True)
load_ws = load_wb['san']

for r in range(0,1506):
    row_value = ""
    for c in range(1,6):
        if load_ws.cell(row=r, column=c).value == None:
            row_value += ""
        else:
            row_value += str(load_ws.cell(row=r, column=c).value)+" "
    all_values.append(row_value)

f = open("상호명충북.txt", 'w', encoding='utf-8')
for seoul_gu in all_values:
    elem.clear()
    elem.send_keys(seoul_gu)

    # 장소 더보기 누르기
    try: # 5개 이상
        driver.find_element_by_css_selector("#info\.search\.place\.more").send_keys(Keys.ENTER)
        #기다려주기
    except: # 5개 이하
        shops = driver.find_elements_by_class_name("PlaceItem")
        for shop in shops:
            # 이름
            name = shop.find_element_by_css_selector("div.head_item.clickArea > strong > a.link_name").text
            # 별 리뷰
            try:
                rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                rating_num = rating[:-1]
                # 참조 리뷰
                review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                if int(rating_num)+int(review) > 1:
                    f.writelines([seoul_gu+",", name+",", rating_num+",", review+"\n"])
            except:
                #f.writelines([seoul_gu+",", name+",", "0", "0\n"])
                continue

            #ws.append([seoul_gu, name])
        continue

    # 1 페이지 돌아가기
    try:
        # time.sleep(0.5)
        # driver.find_element_by_id("info.search.page.no1").send_keys(Keys.ENTER)
        # 크롤링 할 페이지 개수 확인하기
        entire = driver.find_element_by_css_selector("#info\.search\.place\.cnt").text.replace(',', '')
        entire_int = int(entire)
        pages = divmod(entire_int, 15)
        entire_page = pages[0] + 1
        if entire_page > 35:
            # 카카오맵은 검색시 34페이지 제한이 존재한다. 또한 python range는 마지막수 - 1 로 검색한다.
            entire_page = 35

        print(seoul_gu, entire_page)

        # 1 페이지 확인
        page_bar = driver.find_element_by_css_selector("#info\.search\.page > div")
        now_page_bar = page_bar.find_element_by_class_name("ACTIVE").text
        # print(now_page_bar)
        now_page = int(now_page_bar)
        # print(type(now_page_bar))

        # 페이지 자동 넘기기
        for n in range(now_page, entire_page):

            print("Now Crawling Page", n)
            # driver.find_element_by_id("info.search.page.no" + page_number).send_keys(Keys.ENTER)
            # driver.find_element_by_id("info.search.page.next").send_keys(Keys.ENTER)
            time.sleep(0.5)
            shops = driver.find_elements_by_class_name("PlaceItem")
            for shop in shops:
                # 이름
                name = shop.find_element_by_css_selector("div.head_item.clickArea > strong > a.link_name").text
                try:
                    rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                    rating_num = rating[:-1]
                    # 참조 리뷰
                    review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                    if int(rating_num) + int(review) > 1:
                        f.writelines([seoul_gu + ",", name + ",", rating_num + ",", review + "\n"])

                except:
                    f.writelines([seoul_gu + ",", name + ",", "0", "0\n"])
                    continue
                # ws.append([seoul_gu, name])

            if n == 34:
                break
            if n % 5 == 0:
                driver.find_element_by_id("info.search.page.next").send_keys(Keys.ENTER)
            elif n > 5:
                page_number = str((n % 5) + 1)
                try:
                    driver.find_element_by_id("info.search.page.no" + page_number).send_keys(Keys.ENTER)
                except:
                    continue
            else:
                page_number = str(n + 1)
                driver.find_element_by_id("info.search.page.no" + page_number).send_keys(Keys.ENTER)
    except: #페이지 하나밖에 없으면
        time.sleep(0.5)
        shops = driver.find_elements_by_class_name("PlaceItem")

        for shop in shops:
            # 이름
            name = shop.find_element_by_css_selector("div.head_item.clickArea > strong > a.link_name").text
            try:
                rating = shop.find_element_by_css_selector("div.rating.clickArea > span.score > a").text
                rating_num = rating[:-1]
                # 참조 리뷰
                review = shop.find_element_by_css_selector("div.rating.clickArea > a > em").text
                if int(rating_num) + int(review) > 1:
                    f.writelines([seoul_gu + ",", name + ",", rating_num + ",", review + "\n"])

            except:
                f.writelines([seoul_gu + ",", name + ",", "0", "0\n"])
                continue
            # ws.append([seoul_gu, name])




print("크롤링 끝!")
#wb.save(filename="엑셀표.xlsx")